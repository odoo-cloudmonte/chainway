from collections import defaultdict
from io import BytesIO
from plistlib import InvalidFileException

from openpyxl import load_workbook

from odoo import models, fields, api
from odoo.exceptions import UserError, ValidationError

class HelpdeskSr(models.Model):
    _name = "helpdesk.sr"
    _rec_name ='device_sn'

    device_sn = fields.Char(string="Device SN")
    description = fields.Char(string="Description")
    model_mo = fields.Char(string="Model No")
    device_condition = fields.Selection([
        ('poor', 'Poor'),
        ('moderate', 'Moderate'),
        ('good','Good')
    ], string="Device Condition", tracking=True)
    remark = fields.Char(string="Remarks")
    

    ticket_id = fields.Many2one(
                'ticket.helpdesk',
                string="Tickets"
            )
    sequence = fields.Integer(string="S.No.", readonly=True)

    @api.model_create_multi
    def create(self, vals_list):
        ticket_sequences = defaultdict(int)

        # Initialize with existing max sequence for each ticket
        for vals in vals_list:
            ticket_id = vals.get("ticket_id")
            if ticket_id and ticket_id not in ticket_sequences:
                last = self.search(
                    [("ticket_id", "=", ticket_id)],
                    order="sequence desc",
                    limit=1,
                )
                ticket_sequences[ticket_id] = last.sequence or 0

        # Assign sequence
        for vals in vals_list:
            ticket_id = vals.get("ticket_id")
            if ticket_id and not vals.get("sequence"):
                ticket_sequences[ticket_id] += 1
                vals["sequence"] = ticket_sequences[ticket_id]

        return super().create(vals_list)
    
    
    

from odoo import models, fields
import base64

class HelpdeskSrImportWizard(models.TransientModel):
    _name = "helpdesk.sr.import.wizard"
    _description = "Import Service Requests"

    file = fields.Binary(required=True)
    filename = fields.Char()
    ticket_id = fields.Many2one(
        "ticket.helpdesk",
        required=True,
    )


    def action_sr_import(self):
        self.ensure_one()

        if not self.file:
            raise ValidationError("Please upload an Excel (.xlsx) file.")

        # Validate file extension
        if not self.filename or not self.filename.lower().endswith(".xlsx"):
            raise ValidationError(
                "Invalid file format.\nPlease upload only an Excel (.xlsx) file."
            )

        # Validate workbook
        try:
            workbook = load_workbook(
                filename=BytesIO(base64.b64decode(self.file)),
                data_only=True
            )
        except (InvalidFileException, Exception):
            raise ValidationError(
                "The uploaded file is not a valid Excel (.xlsx) file."
            )

        sheet = workbook.active

        # Validate headers
        expected_headers = [
            "Device SN",
            "Description",
            "Model No",
            "Device Condition",
            "Remark",
        ]

        headers = [cell.value for cell in sheet[1]]

        if headers != expected_headers:
            raise ValidationError(
                "Invalid import format.\n"
                "Please download and use the provided import template."
            )

        sequence = len(self.ticket_id.sr_ids) + 1

        valid_conditions = {"good", "moderate", "poor"}

        for row_no, row in enumerate(
            sheet.iter_rows(min_row=2, values_only=True),
            start=2,
        ):

            device_sn = row[0]
            description = row[1]
            model_mo = row[2]
            device_condition = (row[3] or "").strip().lower()
            remark = row[4]

            if device_condition and device_condition not in valid_conditions:
                raise ValidationError(
                    f"Invalid Device Condition '{row[3]}' at row {row_no}.\n"
                    "Allowed values are: good, moderate, poor."
                )

            self.env["helpdesk.sr"].create({
                "ticket_id": self.ticket_id.id,
                "device_sn": device_sn,
                "description": description,
                "model_mo": model_mo,
                "device_condition": device_condition,
                "remark": remark,
                "sequence": sequence,
            })

            sequence += 1

        return {"type": "ir.actions.act_window_close"}