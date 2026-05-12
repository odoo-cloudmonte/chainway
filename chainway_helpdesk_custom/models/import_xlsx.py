# import base64
# import io
# import openpyxl
# import requests
# from odoo import models, fields
# from datetime import datetime

# class DeviceImportWizard(models.TransientModel):
#     _name = 'device.import.wizard'
#     _description = 'Import Devices'

#     file = fields.Binary(string="Upload Excel", required=True)
#     file_name = fields.Char()

#     def _parse_date(self, value):
#         """Convert Excel date safely"""
#         if not value:
#             return False
#         if isinstance(value, datetime):
#             return value.date()
#         return value  # already string or valid

#     def action_import(self):
#         wb = openpyxl.load_workbook(io.BytesIO(base64.b64decode(self.file)),data_only=True)
#         sheet = wb.active

#         headers = [cell.value for cell in sheet[1]]

#         for index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
#             row_dict = dict(zip(headers, row))

#             image_data = row_dict.get('POD Image')

#         # for index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):

#             try:
#                 # 🚫 Skip completely empty row
#                 if not any(row):
#                     continue

#                 device_sn = row[0]

#                 # 🚫 Skip if no device_sn (optional rule)
#                 if not device_sn:
#                     continue

#                 # 🔁 Prevent duplicate crash
#                 existing = self.env['device.inventory'].search([
#                     ('device_sn', '=ilike', device_sn)
#                 ], limit=1)
#                 if existing:
#                     continue  # skip instead of crash

#                 # 🔍 Partner handling
#                 partner = False
#                 if row[15]:
#                     partner = self.env['res.partner'].search(
#                         [('name', '=', row[14])], limit=1
#                     )
#                     if not partner:
#                         partner = self.env['res.partner'].create({
#                             'name': row[14]
#                         })

#                 # 🖼 Image handling
#                 pod_copy = False
#                 # image_data = row[30]

#                 if image_data:
#                     if isinstance(image_data, str) and image_data.startswith('http'):
#                         try:
#                             response = requests.get(image_data, timeout=5)
#                             if response.status_code == 200:
#                                 pod_copy = base64.b64encode(response.content)
#                         except Exception:
#                             pod_copy = False
#                     else:
#                         try:    
#                             pod_copy = image_data if isinstance(image_data, bytes) else base64.b64decode(image_data)
#                         except Exception:
#                             pod_copy = False

#                 # ✅ Create record
#                 self.env['device.inventory'].create({
#                     'device_sn': device_sn,
#                     # 'accessories_sn': row[1],
#                     'type': (row[1] or '').lower(),
#                     'bt_mac': row[2],
#                     'imei1': row[3],
#                     'imei2': row[4],
#                     'wifi_mac': row[5],
#                     'model': row[6],
#                     'description': row[7],
#                     'sw_version': row[8],
#                     'configuration_in': row[9],
#                     'configuration_out': row[10],
#                     'warranty_wef': self._parse_date(row[11]),
#                     'warranty_category': row[12],
#                     'warranty_upto': self._parse_date(row[13]),
#                     'end_user_name': partner.id if partner else False,
#                     'po_date': self._parse_date(row[15]),
#                     'po_no': row[16],
#                     'invoice_no': row[17],
#                     'invoice_date': self._parse_date(row[18]),
#                     'location': row[19],
#                     'location_code': row[20],
#                     'shipping_address_po': row[21],
#                     'shipping_address_invoice': row[22],
#                     'delivery_location': row[23],
#                     'courier_name': row[24],
#                     'tracking_id': row[25],
#                     'delivery_date': self._parse_date(row[26]),
#                     'chainway_reference': row[27],
#                     'remark': row[28],
#                     'pod_copy': pod_copy,
#                 })

#             except Exception as e:
#                 raise Exception(f"Error on Excel row {index}: {str(e)}")


import base64
import io
import openpyxl
import requests

from odoo import models, fields
from odoo.exceptions import ValidationError
from datetime import datetime


class DeviceImportWizard(models.TransientModel):
    _name = 'device.import.wizard'
    _description = 'Import Devices'

    file = fields.Binary(string="Upload Excel", required=True)
    file_name = fields.Char()

    # ------------------------
    # HELPERS
    # ------------------------

    def _parse_date(self, value, field_name, row_index):
        if not value:
            return False

        if isinstance(value, datetime):
            return value.date()

        if isinstance(value, str):
            value = value.strip()

            # ❌ Excel formula
            if value.startswith('='):
                raise ValidationError(
                    f"Row {row_index}\n"
                    f"Column: {field_name}\n"
                    f"Issue: Excel formula detected ({value}). Use actual date."
                )

            # Try formats
            for fmt in ['%Y-%m-%d', '%d-%m-%Y', '%d/%m/%Y']:
                try:
                    return datetime.strptime(value, fmt).date()
                except Exception:
                    continue

        raise ValidationError(
            f"Row {row_index}\n"
            f"Column: {field_name}\n"
            f"Issue: Invalid date '{value}'"
        )

    def _validate_selection(self, value, allowed, field_name, row_index):
        if not value:
            return False

        value = str(value).strip().lower()

        if value not in allowed:
            raise ValidationError(
                f"Row {row_index}\n"
                f"Column: {field_name}\n"
                f"Issue: Invalid value '{value}'\n"
                f"Allowed: {', '.join(allowed)}"
            )

        return value

    # ------------------------
    # MAIN IMPORT
    # ------------------------

    # def action_import(self):
    #     wb = openpyxl.load_workbook(
    #         io.BytesIO(base64.b64decode(self.file)),
    #         data_only=True
    #     )
    #     sheet = wb.active

    #     headers = [cell.value for cell in sheet[1]]

    #     TYPE_VALUES = ['device', 'accessory', 'accessories']
    #     WARRANTY_VALUES = ['standard', 'extended', 'comprehensive', 'none']

    #     for index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
    #         row_dict = dict(zip(headers, row))
    #         image_data = row_dict.get('POD Image')

    #         try:
    #             # 🚫 Skip empty row
    #             if not any(row):
    #                 continue

    #             # ------------------------
    #             # REQUIRED FIELD
    #             # ------------------------
    #             device_sn = row_dict.get('Device SN')
    #             if not device_sn:
    #                 raise ValidationError(
    #                     f"Row {index}\nColumn: Device SN\nIssue: Value is required"
    #                 )

    #             # Duplicate check
    #             existing = self.env['device.inventory'].search([
    #                 ('device_sn', '=ilike', device_sn)
    #             ], limit=1)

    #             if existing:
    #                 continue

    #             # ------------------------
    #             # PARTNER
    #             # ------------------------
    #             partner = False
    #             partner_name = row_dict.get('End User')

    #             if partner_name:
    #                 partner = self.env['res.partner'].search(
    #                     [('name', '=', partner_name)], limit=1
    #                 )
    #                 if not partner:
    #                     partner = self.env['res.partner'].create({
    #                         'name': partner_name
    #                     })

    #             # ------------------------
    #             # IMAGE
    #             # ------------------------
    #             pod_copy = False

    #             if image_data:
    #                 if isinstance(image_data, str) and image_data.startswith('http'):
    #                     try:
    #                         response = requests.get(image_data, timeout=5)
    #                         if response.status_code == 200:
    #                             pod_copy = base64.b64encode(response.content)
    #                     except Exception:
    #                         pod_copy = False
    #                 else:
    #                     try:
    #                         pod_copy = image_data if isinstance(image_data, bytes) else base64.b64decode(image_data)
    #                     except Exception:
    #                         pod_copy = False

    #             # ------------------------
    #             # CREATE RECORD
    #             # ------------------------
    #             self.env['device.inventory'].create({

    #                 'device_sn': device_sn,

    #                 'type': self._validate_selection(
    #                     row_dict.get('Type'),
    #                     TYPE_VALUES,
    #                     'Type',
    #                     index
    #                 ),

    #                 'bt_mac': row_dict.get('BT MAC'),
    #                 'imei1': row_dict.get('IMEI1'),
    #                 'imei2': row_dict.get('IMEI2'),
    #                 'wifi_mac': row_dict.get('WiFi MAC'),
    #                 'model': row_dict.get('Model'),
    #                 'description': row_dict.get('Description'),
    #                 'sw_version': row_dict.get('SW Version'),
    #                 'configuration_in': row_dict.get('Configuration IN'),
    #                 'configuration_out': row_dict.get('Configuration OUT'),

    #                 'warranty_wef': self._parse_date(
    #                     row_dict.get('Warranty WEF'),
    #                     'Warranty WEF',
    #                     index
    #                 ),

    #                 'warranty_category': self._validate_selection(
    #                     row_dict.get('Warranty Category'),
    #                     WARRANTY_VALUES,
    #                     'Warranty Category',
    #                     index
    #                 ),

    #                 'warranty_upto': self._parse_date(
    #                     row_dict.get('Warranty Upto'),
    #                     'Warranty Upto',
    #                     index
    #                 ),

    #                 'end_user_name': partner.id if partner else False,

    #                 'po_date': self._parse_date(
    #                     row_dict.get('PO Date'),
    #                     'PO Date',
    #                     index
    #                 ),

    #                 'po_no': row_dict.get('PO No'),
    #                 'invoice_no': row_dict.get('Invoice No'),

    #                 'invoice_date': self._parse_date(
    #                     row_dict.get('Invoice Date'),
    #                     'Invoice Date',
    #                     index
    #                 ),

    #                 'location': row_dict.get('Location'),
    #                 'location_code': row_dict.get('Location Code'),

    #                 'shipping_address_po': row_dict.get('Shipping Address PO'),
    #                 'shipping_address_invoice': row_dict.get('Shipping Address Invoice'),
    #                 'delivery_location': row_dict.get('Confirm Delivery Location'),

    #                 'courier_name': row_dict.get('Courier Name'),
    #                 'tracking_id': row_dict.get('Tracking ID'),

    #                 'delivery_date': self._parse_date(
    #                     row_dict.get('Delivery Date'),
    #                     'Delivery Date',
    #                     index
    #                 ),

    #                 'chainway_reference': row_dict.get('Chainway Reference'),
    #                 'remark': row_dict.get('Remark'),
    #                 'pod_copy': pod_copy,
    #             })

    #         except ValidationError:
    #             raise  # show clean popup
    #         except Exception as e:
    #             raise ValidationError(
    #                 f"Error on Excel row {index}\n{str(e)}"
    #             )
            

    def action_import(self):
        wb = openpyxl.load_workbook(
            io.BytesIO(base64.b64decode(self.file)),
            data_only=True
        )

        sheet = wb.active

        headers = [cell.value for cell in sheet[1]]

        TYPE_VALUES = ['device', 'accessory', 'accessories']
        WARRANTY_VALUES = ['standard', 'extended', 'comprehensive', 'none']

        for index, row in enumerate(
            sheet.iter_rows(min_row=2, values_only=True),
            start=2
        ):

            row_dict = dict(zip(headers, row))
            image_data = row_dict.get('POD Image')

            try:

                # -------------------------------------------------
                # SKIP EMPTY ROW
                # -------------------------------------------------
                if not any(row):
                    continue

                # -------------------------------------------------
                # REQUIRED FIELD
                # -------------------------------------------------
                device_sn = row_dict.get('Device SN')

                if not device_sn:
                    raise ValidationError(
                        f"Row {index}\n"
                        f"Column: Device SN\n"
                        f"Issue: Value is required"
                    )

                # -------------------------------------------------
                # CHECK EXISTING RECORD
                # -------------------------------------------------
                existing = self.env['device.inventory'].search([
                    ('device_sn', '=ilike', device_sn)
                ], limit=1)

                # -------------------------------------------------
                # PARTNER
                # -------------------------------------------------
                partner = False

                partner_name = row_dict.get('End User')

                if partner_name:
                    partner = self.env['res.partner'].search([
                        ('name', '=', partner_name)
                    ], limit=1)

                    if not partner:
                        partner = self.env['res.partner'].create({
                            'name': partner_name
                        })

                # -------------------------------------------------
                # IMAGE
                # -------------------------------------------------
                pod_copy = False

                if image_data:

                    # IMAGE URL
                    if isinstance(image_data, str) and image_data.startswith('http'):

                        try:
                            response = requests.get(image_data, timeout=5)

                            if response.status_code == 200:
                                pod_copy = base64.b64encode(response.content)

                        except Exception:
                            pod_copy = False

                    # BASE64 IMAGE
                    else:
                        try:

                            if isinstance(image_data, bytes):
                                pod_copy = image_data
                            else:
                                pod_copy = base64.b64decode(image_data)

                        except Exception:
                            pod_copy = False

                # -------------------------------------------------
                # VALUES
                # -------------------------------------------------
                vals = {

                    'device_sn': device_sn,

                    'type': self._validate_selection(
                        row_dict.get('Type'),
                        TYPE_VALUES,
                        'Type',
                        index
                    ),

                    'bt_mac': row_dict.get('BT MAC'),
                    'imei1': row_dict.get('IMEI1'),
                    'imei2': row_dict.get('IMEI2'),
                    'wifi_mac': row_dict.get('WiFi MAC'),

                    'model': row_dict.get('Model'),
                    'description': row_dict.get('Description'),
                    'sw_version': row_dict.get('SW Version'),

                    'configuration_in': row_dict.get('Configuration IN'),
                    'configuration_out': row_dict.get('Configuration OUT'),

                    'warranty_wef': self._parse_date(
                        row_dict.get('Warranty WEF'),
                        'Warranty WEF',
                        index
                    ),

                    'warranty_category': self._validate_selection(
                        row_dict.get('Warranty Category'),
                        WARRANTY_VALUES,
                        'Warranty Category',
                        index
                    ),

                    'warranty_upto': self._parse_date(
                        row_dict.get('Warranty Upto'),
                        'Warranty Upto',
                        index
                    ),

                    'end_user_name': partner.id if partner else False,

                    'po_date': self._parse_date(
                        row_dict.get('PO Date'),
                        'PO Date',
                        index
                    ),

                    'po_no': row_dict.get('PO No'),

                    'invoice_no': row_dict.get('Invoice No'),

                    'invoice_date': self._parse_date(
                        row_dict.get('Invoice Date'),
                        'Invoice Date',
                        index
                    ),

                    'location': row_dict.get('Location'),
                    'location_code': row_dict.get('Location Code'),

                    'shipping_address_po': row_dict.get('Shipping Address PO'),

                    'shipping_address_invoice': row_dict.get(
                        'Shipping Address Invoice'
                    ),

                    'delivery_location': row_dict.get(
                        'Confirm Delivery Location'
                    ),

                    'courier_name': row_dict.get('Courier Name'),
                    'tracking_id': row_dict.get('Tracking ID'),

                    'delivery_date': self._parse_date(
                        row_dict.get('Delivery Date'),
                        'Delivery Date',
                        index
                    ),

                    'chainway_reference': row_dict.get(
                        'Chainway Reference'
                    ),

                    'remark': row_dict.get('Remark'),

                    'pod_copy': pod_copy,
                }

                # -------------------------------------------------
                # UPDATE OR CREATE
                # -------------------------------------------------
                if existing:
                    existing.write(vals)
                else:
                    self.env['device.inventory'].create(vals)

            except ValidationError:
                raise

            except Exception as e:
                raise ValidationError(
                    f"Error on Excel row {index}\n{str(e)}"
                )