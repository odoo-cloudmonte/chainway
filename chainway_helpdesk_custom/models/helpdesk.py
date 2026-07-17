from datetime import timedelta
import re
import base64
from io import BytesIO
from openpyxl import load_workbook
import xlsxwriter

from odoo import api, models, fields
from odoo.exceptions import UserError, ValidationError

class TicketHelpdesk(models.Model):
    _inherit = 'ticket.helpdesk'

    # device_ids = fields.Many2many(
    #     'device.inventory',
    #     'ticket_device_rel',   # relation table name
    #     'ticket_id',           # column for this model
    #     'device_id',           # column for device model
    #     string="Devices"
    # )

    sr_ids = fields.One2many(
        'helpdesk.sr',   # target model
        'ticket_id',     # inverse field
        string="Service Requests"
    )
    export_file = fields.Binary(string="Export File")
    export_filename = fields.Char(string="File Name")

    sr_device_sn = fields.Char(
        string="Device SN",
        compute="_compute_sr_device_sn",
        store=True
    )

    sr_device_model = fields.Char(
        string="Model No",
        compute="_compute_sr_device_model",
        store=True
    )

    company_name = fields.Char(string="Company Name")
    company_address = fields.Char(string="Company Address")
    contact = fields.Char(string="Contact")
    store = fields.Char(string="Pin Code")
    city = fields.Char(string="City")

    is_overdue = fields.Boolean()
    is_super_overdue = fields.Boolean()

    is_closed = fields.Boolean()
    is_on_time = fields.Boolean()
    # device_condition = fields.Selection([
    #     ('poor', 'Poor'),
    #     ('moderate', 'Moderate'),
    #     ('good','Good')
    # ], string="Device Condition", tracking=True)

    # @api.depends('create_date', 'stage_id')
    # def _compute_states(self):
    #     now = fields.Datetime.now()
    #     for rec in self:
    #         rec.is_overdue = False
    #         rec.is_closed = False
    #         rec.is_on_time = False

    #         if not rec.create_date:
    #             continue

    #         is_closed = rec.stage_id.closing_stage if rec.stage_id else False

    #         if is_closed:
    #             rec.is_closed = True
    #         elif now > rec.create_date + timedelta(hours=48):
    #             rec.is_overdue = True
    #         else:
    #             rec.is_on_time = True
    
    # def action_export_sr(self):
    #     self.ensure_one()

    #     output = BytesIO()
    #     workbook = xlsxwriter.Workbook(output)
    #     sheet = workbook.add_worksheet('Service Requests')

    #     # Header format
    #     header_format = workbook.add_format({
    #         'bold': True,
    #         'bg_color': '#D3D3D3',
    #         'border': 1
    #     })

    #     # Data format
    #     data_format = workbook.add_format({
    #         'border': 1
    #     })

    #     headers = [
    #         'Device SN',
    #         'Description',
    #         'Model No',
    #         'Device Condition',
    #         'Remarks'
    #     ]

    #     # Write headers
    #     for col, header in enumerate(headers):
    #         sheet.write(0, col, header, header_format)

    #     row = 1

    #     for line in self.sr_ids:
    #         sheet.write(row, 0, line.device_sn or '', data_format)
    #         sheet.write(row, 1, line.description or '', data_format)
    #         sheet.write(row, 2, line.model_mo or '', data_format)
    #         sheet.write(
    #             row,
    #             3,
    #             dict(line._fields['device_condition'].selection).get(
    #                 line.device_condition,
    #                 ''
    #             ),
    #             data_format
    #         )
    #         sheet.write(row, 4, line.remark or '', data_format)

    #         row += 1

    #     workbook.close()
    #     output.seek(0)

    #     file_data = base64.b64encode(output.read())

    #     self.write({
    #         'export_file': file_data,
    #         'export_filename': f'Service_Request_{self.id}.xlsx'
    #     })

    #     return {
    #         'type': 'ir.actions.act_url',
    #         'url': (
    #             '/web/content?model=ticket.helpdesk'
    #             '&id=%s'
    #             '&field=export_file'
    #             '&filename_field=export_filename'
    #             '&download=true'
    #         ) % self.id,
    #         'target': 'self',
    #     }
    
    def action_export_sr(self):
        output = BytesIO()
        workbook = xlsxwriter.Workbook(output)
        sheet = workbook.add_worksheet('Service Requests')

        # Formats
        header_format = workbook.add_format({
            'bold': True,
            'align': 'center',
            'valign': 'vcenter',
            'border': 1,
            'bg_color': '#D9EAD3',
        })

        cell_format = workbook.add_format({
            'align': 'left',
            'valign': 'vcenter',
            'border': 1,
            'text_wrap': True,
        })

        date_format = workbook.add_format({
            'align': 'center',
            'valign': 'vcenter',
            'border': 1,
            'num_format': 'yyyy-mm-dd hh:mm:ss',
        })

        headers = [
            'Service Requests/Tickets',
            'Service Requests/Device SN',
            'Service Requests/Description',
            'Service Requests/Model No',
            'Service Requests/Device Condition',
            'Service Requests/Remarks',
            'Assigned User',
            'Team Leader',
            'Stage',
            'Creation Date',
            'Last Update Date',
        ]

        # Write headers
        for col, header in enumerate(headers):
            sheet.write(0, col, header, header_format)

        # Set column widths
        sheet.set_column('A:A', 25)  # Ticket
        sheet.set_column('B:B', 25)  # Device SN
        sheet.set_column('C:C', 40)  # Description
        sheet.set_column('D:D', 25)  # Model No
        sheet.set_column('E:E', 20)  # Device Condition
        sheet.set_column('F:F', 40)  # Remarks
        sheet.set_column('G:G', 25)  # Assigned User
        sheet.set_column('H:H', 25)  # Team Leader
        sheet.set_column('I:I', 25)  # Stage
        sheet.set_column('J:J', 25)  # Creation Date
        sheet.set_column('K:K', 25)  # Last Updated Date

        row = 1

        for ticket in self:
            for line in ticket.sr_ids:
                sheet.write(row, 0, ticket.display_name or '', cell_format)
                sheet.write(row, 1, line.device_sn or '', cell_format)
                sheet.write(row, 2, line.description or '', cell_format)
                sheet.write(row, 3, line.model_mo or '', cell_format)

                sheet.write(
                    row,
                    4,
                    dict(line._fields['device_condition'].selection).get(
                        line.device_condition, ''
                    ),
                    cell_format,
                )

                sheet.write(row, 5, line.remark or '', cell_format)
                sheet.write(
                    row,
                    6,
                    ticket.assigned_user_id.name if ticket.assigned_user_id else '',
                    cell_format,
                )
                sheet.write(
                    row,
                    7,
                    ticket.team_head_id.name if ticket.team_head_id else '',
                    cell_format,
                )
                sheet.write(row, 8, ticket.stage_id.name or '', cell_format)
                

                if ticket.create_date:
                    sheet.write_datetime(
                        row,
                        9,
                        ticket.create_date,
                        date_format,
                    )
                else:
                    sheet.write(row, 9, '', cell_format)
                    
                if ticket.last_update_date:
                    sheet.write_datetime(
                        row,
                        10,
                        ticket.last_update_date,
                        date_format,
                    )
                else:
                    sheet.write(row, 10, '', cell_format)

                row += 1

        # Freeze header row
        sheet.freeze_panes(1, 0)

        workbook.close()
        output.seek(0)

        file_data = base64.b64encode(output.read())

        attachment = self.env['ir.attachment'].create({
            'name': 'service_requests.xlsx',
            'type': 'binary',
            'datas': file_data,
            'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        })

        return {
            'type': 'ir.actions.act_url',
            'url': '/web/content/%s?download=true' % attachment.id,
            'target': 'self',
        }

    @api.depends('sr_ids.device_sn')
    def _compute_sr_device_sn(self):
        for rec in self:
            serials = rec.sr_ids.mapped('device_sn')

            # Remove False/empty values
            serials = [sn for sn in serials if sn]

            rec.sr_device_sn = ', '.join(serials)

    @api.depends('sr_ids.model_mo')
    def _compute_sr_device_model(self):
        for rec in self:
            serials = rec.sr_ids.mapped('model_mo')

            # Remove False/empty values
            serials = [sn for sn in serials if sn]

            rec.sr_device_model = ', '.join(serials)


    def cron_update_ticket_states(self):
        now = fields.Datetime.now()

        # tickets = self.search([])
        tickets = self.search([
            ('stage_id.closing_stage', '=', False)
        ])

        for rec in tickets:
            # Reset all states
            rec.is_overdue = False
            rec.is_closed = False
            rec.is_on_time = False
            rec.is_super_overdue = False

            if not rec.create_date:
                continue

            is_closed = rec.stage_id.closing_stage if rec.stage_id else False

            # 🟢 CLOSED
            if is_closed:
                rec.is_closed = True
                rec.is_on_time = False
                rec.is_overdue = False
                rec.is_super_overdue = False

            elif now > rec.create_date + timedelta(hours=72):
                rec.is_overdue = False
                rec.is_closed = False
                rec.is_on_time = False
                rec.is_super_overdue = True

            # 🔴 OVERDUE
            elif now > rec.create_date + timedelta(hours=48):
                rec.is_overdue = True
                rec.is_closed = False
                rec.is_on_time = False
                rec.is_super_overdue = False

            

            # 🟡 ON TIME
            else:
                rec.is_on_time = True
                rec.is_overdue = False
                rec.is_closed = False
                rec.is_super_overdue = False


    @api.constrains('create_date', 'start_date', 'end_date')
    def _check_ticket_dates(self):
        for rec in self:

            # create_date <= start_date
            if rec.create_date and rec.start_date:
                if rec.start_date < rec.create_date:
                    raise ValidationError(
                        "Start Date cannot be earlier than Creation Date."
                    )

            # start_date <= end_date
            if rec.start_date and rec.end_date:
                if rec.end_date < rec.start_date:
                    raise ValidationError(
                        "End Date cannot be earlier than Start Date."
                    )

            # create_date <= end_date
            if rec.create_date and rec.end_date:
                if rec.end_date < rec.create_date:
                    raise ValidationError(
                        "End Date cannot be earlier than Creation Date."
                    )
                
    @api.constrains('phone')
    def _check_phone_number(self):
        for rec in self:
            if rec.phone:
                phone = rec.phone.strip()

                # Allow only digits, optional + at start, length 10–15
                if not re.match(r'^\+?\d{10,15}$', phone):
                    raise ValidationError(
                        "Phone number must be 10–15 digits and can optionally start with '+'."
                    )
                    
    
    
    def action_import_sr(self):
        self.ensure_one()

        return {
            "type": "ir.actions.act_window",
            "name": "Import Service Requests",
            "res_model": "helpdesk.sr.import.wizard",
            "view_mode": "form",
            "target": "new",
            "context": {
                "default_ticket_id": self.id,
            },
        }