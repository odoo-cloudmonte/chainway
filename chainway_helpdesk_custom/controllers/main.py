import base64
import io

from odoo import http
from odoo.exceptions import ValidationError
from odoo.http import request
from datetime import date, datetime

class WarrantyController(http.Controller):

    @http.route(['/warranty-check'], type='http', auth='public', website=True)
    def warranty_form(self, **kwargs):
        return request.render('chainway_helpdesk_custom.warranty_form_template')

    @http.route(['/warranty-result'], type='http', auth='public', website=True, methods=['POST'])
    def warranty_result(self, **post):
        search_sn = post.get('search_sn')

        device = request.env['device.inventory'].sudo().search([
            '|', '|',
            ('device_sn', '=', search_sn),
            ('imei1', '=', search_sn),
            ('imei2', '=', search_sn),
        ], limit=1)

        result = {}
        if device:
            if device.warranty_upto:
                status = "In Warranty" if device.warranty_upto >= date.today() else "Expired"
            else:
                status = "No Warranty Info"

            result = {
                'found': True,
                'device': device,
                'status': status,
            }
        else:
            result = {
                'found': False,
                'status': "Device Not Found"
            }

        return request.render('chainway_helpdesk_custom.warranty_result_template', result)

    @http.route('/smart/form', type='http', auth='public', website=True)
    def smart_form(self, **kw):

        return request.render(
            'chainway_helpdesk_custom.smart_form_template',
            {
                'data': kw
            }
        )

    @http.route('/my/helpdesk/form', type='http', auth='user', website=True)
    def helpdesk_form(self, **kw):

        return request.render('chainway_helpdesk_custom.portal_helpdesk_form', {
            'page_name': 'helpdesk_form'
        })
    
    @http.route('/my/helpdesk/submit', type='http', auth='public', website=True, csrf=True)
    def submit_ticket(self, **post):

        

        models = request.httprequest.form.getlist('model_number[]')
        serials = request.httprequest.form.getlist('serial_number[]')
        descs = request.httprequest.form.getlist('description[]')

        sr_lines = []

        max_len = max(len(models), len(serials), len(descs))

        for i in range(max_len):
            model = (models[i] if i < len(models) else '').strip()
            serial = (serials[i] if i < len(serials) else '').strip()
            desc = (descs[i] if i < len(descs) else '').strip()

            # ✅ Always create row (even if empty)
            sr_lines.append((0, 0, {
                'model_mo': model or False,
                'device_sn': serial or False,
                'description': desc or False,
            }))

        # ticket = request.env['ticket.helpdesk'].sudo().create({
        #     'customer_name': post.get('company_name'),
        #     'subject': post.get('problem_type'),
        #     'description': post.get('description'),
        #     'email': post.get('email'),
        #     'phone': post.get('contact_number'),
        # })

        # models = request.httprequest.form.getlist('model_number[]')
        # serials = request.httprequest.form.getlist('serial_number[]')
        # descs = request.httprequest.form.getlist('device_description[]')

        # for model, serial, desc in zip(models, serials, descs):
        #     if model or serial or desc:
        #         request.env['helpdesk.sr'].sudo().create({
        #             'ticket_id': ticket.id,
        #             'model_mo': model,
        #             'device_sn': serial,
        #             'description': desc,
        #         })
        try:
        
            stage = request.env['ticket.stage'].sudo().search([('name', '=', 'New')], limit=1)
            ticket = request.env['ticket.helpdesk'].sudo().create({
                'customer_id':request.env.user.partner_id.id,
                'company_name': post.get('company_name'),
                'company_address':post.get('company_address'),
                'contact': post.get('contact_person'),
                'store':post.get('store_location'),
                # 'customer_name': post.get('company_name'),
                'subject': post.get('problem_type'),
                'description': post.get('description'),
                'email': post.get('email'),
                'phone': post.get('contact_number'),
                'sr_ids': sr_lines,
                'stage_id':stage.id,
            })

            

            files = request.httprequest.files.getlist('attachment')

            for file in files:
                if file:
                    request.env['ir.attachment'].sudo().create({
                        'name': file.filename,
                        'datas': base64.b64encode(file.read()),
                        'res_model': 'ticket.helpdesk',
                        'res_id': ticket.id,
                        'type': 'binary',
                    })

            return request.redirect('/my/tickets')
        
        except ValidationError as e:
            return request.render('chainway_helpdesk_custom.portal_helpdesk_form', {
                'error_message': str(e),
                  # keep old values
            })
    

    @http.route(['/my/devices'], type='http', auth="user", website=True)
    def my_devices(self, **kwargs):

        user = request.env.user

        if not user.show_device_ui:
            return request.redirect('/')
        
        partner = request.env.user.partner_id

        devices = request.env['device.inventory'].sudo().search([
            ('end_user_name', '=', partner.id)
        ])

        return request.render(
            'chainway_helpdesk_custom.portal_device_list',
            {
                'devices': devices
            }
        )

        # if request.env.user.show_device_ui:

        #     return request.render(
        #         'chainway_helpdesk_custom.portal_device_list',
        #         {
        #             'devices': devices
        #         }
        #     )
        
        # return request.redirect('/my/tickets')

    @http.route('/my/devices/download_excel', auth='user', type='http')
    def download_device_excel(self, **kw):
        """Generate and download device data as Excel file"""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        except ImportError:
            return request.not_found()
        
        # Get devices for current user
        device_env = request.env['device.inventory']
        devices = device_env.sudo().search([
            ('end_user_name', '=', request.env.user.partner_id.id)
        ])
        
        # Create workbook
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "Devices"
        
        # Define styles
        header_fill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        border_style = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        cell_alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        
        # Define columns
        columns = [
            ('Sr. No', 8),
            ('Vendor Name', 25),
            ('PO Date', 15),
            ('PO No.', 15),
            ('Invoice No.', 15),
            ('Invoice Date', 15),
            ('Store Name', 20),
            ('Store Code', 15),
            ('Ship. Address (PO)', 25),
            ('Ship. Address (Invoice)', 25),
            ('Delivery Location', 20),
            ('Serial Number', 20),
            ('Description', 25),
            ('Courier', 15),
            ('Tracking ID', 20),
            ('Delivery Date', 15),
            ('POD Status', 12),
        ]
        
        # Write headers
        for col_idx, (header_text, width) in enumerate(columns, 1):
            cell = worksheet.cell(row=1, column=col_idx)
            cell.value = header_text
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = border_style
            worksheet.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width
        
        # Write data rows
        for row_idx, device in enumerate(devices, 2):
            row_data = [
                row_idx - 1,  # Sr. No
                device.end_user_name.name if device.end_user_name else '-',
                str(device.po_date) if device.po_date else '-',
                device.po_no if device.po_no else '-',
                device.invoice_no if device.invoice_no else '-',
                str(device.invoice_date) if device.invoice_date else '-',
                device.location if device.location else '-',
                device.location_code if device.location_code else '-',
                device.shipping_address_po if device.shipping_address_po else '-',
                device.shipping_address_invoice if device.shipping_address_invoice else '-',
                device.delivery_location if device.delivery_location else '-',
                device.device_sn if device.device_sn else '-',
                device.description if device.description else '-',
                device.courier_name if device.courier_name else '-',
                device.tracking_id if device.tracking_id else '-',
                str(device.delivery_date) if device.delivery_date else '-',
                'Yes' if device.pod_copy else 'No',
            ]
            
            for col_idx, value in enumerate(row_data, 1):
                cell = worksheet.cell(row=row_idx, column=col_idx)
                cell.value = value
                cell.alignment = cell_alignment
                cell.border = border_style
        
        # Freeze header row
        worksheet.freeze_panes = "A2"
        
        # Generate file
        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)
        
        # Return file
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"Devices_{timestamp}.xlsx"
        
        return request.make_response(
            output.getvalue(),
            headers=[
                ('Content-Type', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'),
                ('Content-Disposition', f'attachment; filename="{filename}"'),
                ('Content-Length', len(output.getvalue())),
            ]
        )
        

from odoo import http
from odoo.http import request
from odoo.addons.web.controllers.home import Home

class PortalLoginRedirect(Home):

    def _login_redirect(self, uid, redirect=None):
        user = request.env['res.users'].sudo().browse(uid)

        # If portal user → redirect to homepage
        if user.has_group('base.group_portal'):
            return '/'

        # default behavior
        return super()._login_redirect(uid, redirect=redirect)